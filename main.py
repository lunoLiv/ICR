from openpyxl import load_workbook

wb = load_workbook(filename='teste.xlsx')
ws = wb.active

gap = ws['L2':'L15074']

for i in gap:
    for j in i:
        print(j.value)
        input()