### WIP ###
### WIP ###
### WIP ###


from openpyxl import load_workbook
import os, datetime

import warnings
warnings.simplefilter("ignore")

def Pegar_Dados_02(local):
    wb = load_workbook(filename=local, read_only=True)
    
    try:
        ws = wb['BASE']
    except:
        ws = wb.active

    dados = []

    valores = ws[f'A14:AL{ws.max_row}']
    a13 = ws[f'A13:AL13']
    
    cabeca = []
    
    for i in a13:
        for j in i:
            cabeca.append(j.value)
    #input(cabeca)

    counter = 0
    for i in valores:
        
        temp = []
        
        for j in i:
            temp.append(j.value)
        dados.append(temp)
        
        if temp.count(None) == len(temp) and counter == 0:
            return dados, cabeca
            counter += 1    

        if temp.count(None) == len(temp) and counter == 1:
            counter += 1    

        if temp.count(None) == len(temp) and counter == 2:
            return dados, cabeca

    return dados, cabeca

def Limpar_Dados_02(local,caminho):
    dados, header = Pegar_Dados_02(local)
    
    entrada = []
'''
    for i in dados:

        try:
            pasta = i[header.index('PASTA')]
        except:
            pasta = '-'
        
        try:
            sinistro = i[header.index('APOLICE/SINISTRO')]
        except:
            sinistro = '-'

        try:
            aviso = i[header.index('AVISO')]
        except:
            aviso = '-'

        try:
            aluguel = i[header.index('VL.ALUGUEL')]
        except:
            aluguel = '-'

        try:
            tratativa = i[header.index('TRATATIVA')]
        except:
            tratativa = '-'

        try:
            corretor = i[header.index('CORRETOR')]
        except:
            corretor = '-'

        try:
            endereco = i[header.index('ENDEREÇO')]
        except:
            endereco = '-'

        try:
            cidade = i[header.index('CIDADE')]
        except:
            cidade = '-'

        try:
            uf = i[header.index('UF')]
        except:
            uf = '-'

        try:
            imobiliaria = i[header.index('IMOBILIÁRIA')]
        except:
            imobiliaria = '-'

        try:
            endereco = endereco.replace(',',' ')
        except:
            pass
        '''
        entradaDados = [pasta,sinistro,aviso,aluguel,tratativa,corretor,endereco,cidade,uf,imobiliaria,caminho]
        counter = 0
        for i in entradaDados:
            if i == None:
                entradaDados[counter] = '-' 
            counter += 1
        entrada.append(entradaDados)
    return entrada

'''
        pasta = i[0]
        sinistro = i[1]
        aviso = i[2]
        aluguel = 'None'
        tratativa = i[7]
        corretor = i[11]
        endereco = i[15]
        cidade = i[16]
        uf = 'None'
        imobiliaria = i[29]
'''
os.chdir('arquivos')
pasta = [i for i in os.listdir() if i[0][0] == 'P']

#counter = 1
tudo = []
for i in pasta:
    print(i)
    dados = Limpar_Dados_02(i,f'\\192.168.1.54\TokioMarine$\OCUPADO\Supervisão\Planilhas - controles mensais\{i}')
    tudo.append(dados)
os.chdir('C:\\Users\\luiz.leite\\Desktop\\Python\\Excel\\desafio')
with open('a.csv','w') as f:
    f.write('PASTA;SINISTRO;AVISO;VL.ALUGUEL;TRATATIVA;CORRETOR;ENDEREÇO;CIDADE;UF;IMOBILIÁRIA;ORIGEM\n')
    for j in tudo:
        for i in j:

            f.write(f'{i[0]};{i[1]};{i[2]};{i[3]};{i[4]};{i[5]};{i[6]};{i[7]};{i[8]};{i[9]};{i[10]}\n')
    
    
    
    
    #counter += len(i)
    #Entrar_Dados_02('teste.xlsx',dados,f'\\192.168.1.54\TokioMarine$\OCUPADO\Supervisão\Planilhas - controles mensais\{i}',counter)
def Entrar_Dados_02(local,entrada,caminho,counter):
    counter = counter
    for i in entrada:
        wb = load_workbook(filename=local)
        ws = wb.active

        ws[f'A{counter}'] = i[0]
        ws[f'B{counter}'] = i[1]
        ws[f'C{counter}'] = i[2]
        ws[f'D{counter}'] = i[3]
        ws[f'E{counter}'] = i[4]
        ws[f'F{counter}'] = i[5]
        ws[f'G{counter}'] = i[6]
        ws[f'H{counter}'] = i[7]
        ws[f'I{counter}'] = i[8]
        ws[f'J{counter}'] = i[9]
        ws[f'K{counter}'] = caminho
    wb.save(filename='C:\\Users\\luiz.leite\\Desktop\\Python\\Excel\\desafioteste.xlsx')

























'''
def Encontrar_Header(local):
    
    header = ['PASTA','SINISTRO','AVISO','VL.ALUGUEL','TRATATIVA','CORRETOR','ENDEREÇO','CIDADE','UF']

    wb = load_workbook(filename=local, read_only=True)
    ws = wb.active

    for i in ws.iter_rows():
        for cell in i:
            valor = cell.value
            if type(valor) == str:
                if valor.upper() in header:
                    return cell.row
    
def Pegar_Header(local,linha):
    header = []

    wb = load_workbook(filename=local, read_only=True)
    ws = wb.active   

    valores = ws[f'A{linha}:AZ{ws.max_row}']

    for i in valores:
        header.append(i)

    return header

print(Encontrar_Header('Acordo CIA.xlsx'))
print(Pegar_Header('Acordo CIA.xlsx',Encontrar_Header('Acordo CIA.xlsx')))
'''

''''
def Dados_Gerais(local):
  
    wb = load_workbook(filename=local, read_only=True)
    ws = wb.active

    sinistros = ws['C2':'C405']
    casos = []

    for i in sinistros:
        for j in i:
            casos.append(j.value)

    return casos

# Dados TOO AGENDA E TOO AGENDA JURIDICO
def Dados_Agenda(local,sheet,coluna,linhaInicial):
    wb = load_workbook(filename=local,read_only=True)
    ws = wb[f'{sheet}']

    #celulas = ws[f'C12:C{ws.max_row}']
    celulas = ws[f'{coluna}{linhaInicial}:{coluna}{ws.max_row}']
    sinistros = []

    for i in celulas:
        for j in i:
            if j.value != None:
                sinistros.append(j.value)

    return sinistros

def Procurar_Duplicados(lista):
    duplicados = {}
    for i in lista:
        if lista.count(i) > 1:
            duplicados[i] = lista.count(i)

    return duplicados



gerais = Dados_Gerais('AGENDA FORTE 03.02.25.xlsx')
juridicos = Dados_Agenda('TOO AGENDA JURIDICO.xlsx','CONTROLE OPERADORES','B','17')
agenda = Dados_Agenda('TOO AGENDA.xlsx','CONTROLE OPERADORES','B','17')
acordos = Dados_Agenda('ACOMPANHAMENTO ACORDO ATT.xlsx','ACORDOS','C','12')

#print(Procurar_Duplicados(todos))

gerais = set(gerais)
juridicos = set(juridicos)
agenda = set(agenda)
acordos = set(acordos)

todos = agenda | juridicos

diferencaTodos = gerais - todos
diferencaAcordos = diferencaTodos - acordos
ineterAcordos = diferencaTodos & acordos

with open('diffs.txt', 'w') as f:
    f.write(f'TOODOS - MINHAAGENDA ({len(diferencaTodos)}):\n{diferencaTodos}\n\n')
    f.write(f'TOODOS - ACORDOS ({len(diferencaAcordos)}):\n{diferencaAcordos}\n\n')
    f.write(f'TOODOS - ACORDOS ({len(ineterAcordos)}):\n{ineterAcordos}\n\n')

'''
