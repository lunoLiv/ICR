from openpyxl import load_workbook
#import PyPDF2

def Dados_BaseToo(local,apolice):
    # Abrir a BASE da TOO.
    wb = load_workbook(filename=local, read_only=True)
    ws = wb['BASE']

    excelApolice = ws['C8':f'C{ws.max_row}']

    listaApolices = [apolice for i in excelApolice]

    print(listaApolices[123])
    linha = listaApolices.index(apolice)
    print(linha,1004600036789 in listaApolices)

Dados_BaseToo('BASE TOO.xlsm',1004600036789)