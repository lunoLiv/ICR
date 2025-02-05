import os
import openpyxl

valor = "1074600081404" 

def procurar(local, valor):
    for arquivo in os.listdir(local):
        if arquivo.endswith(".xlsx"):  
            file_path = os.path.join(local, arquivo)
            
            try:
                
                wb = openpyxl.load_workbook(file_path, data_only=True,read_only=True)
                found = False
                
                for sheet in wb.sheetnames:
                    sheet_obj = wb[sheet]
                    
                    
                    for row in sheet_obj.iter_rows():
                        for cell in row:
                            if cell.value == valor:
                                print(f"o valor '{valor}' esta aqui: {arquivo}, sheet: {sheet}")
                                found = True
                                break
                        if found:
                            break
                    if found:
                        break
                wb.close()
            
            except Exception as e:
                pass
                print(f"Erro {arquivo}: {e}")
